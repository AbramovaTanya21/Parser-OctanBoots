
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import threading
import queue
import time

import openpyxl
from openpyxl import Workbook,load_workbook
import os

        
ChromedriverPuth = 'G:\\NRU\\SP\\Parsers\\selenium\\chromedriver\\win64\\139.0.7258.66\\chromedriver.exe'
CollectionFilePath = "G:\\NRU\\SP\\Parsers\\OceanBoots\\OceanBoots\\OceanBoots\\Catalog_OceanBoots.xlsx"
file_name = "OceanBoots_Date.xlsx"
file_path = "./" + file_name  

class TabInd:
            NAME = 0
            ARTICLE = 1
            BRAND = 2
            PRICE = 3
            SIZE = 4
            DESCRIPTION = 5
            PHOTO = 6
            LINK = 7 
            
def GettingColltction(driver):
    
    wb = load_workbook(CollectionFilePath)
    for sh in wb.worksheets:
       ws  = wb[sh.title]
       # ws = wb.active 
       LinkPages = [] 
       last_collection = ws[2][0].value        
       for row in range(2, ws.max_row+1):  
            current_collection = ws[row][0].value 
            if current_collection == None: break  
            if last_collection !=  current_collection : 
                 ParsingCollection(driver, LinkPages, last_collection)  
                 LinkPages.clear()
            LinkPages.append(ws[row][2].value) 
            last_collection = current_collection
       ParsingCollection(driver, LinkPages, last_collection)
    driver.quit() 

def ParsingCollection(driver, LinkPages, last_collection):
   print(f'GettingLinks: Начат сбор ссылок для коллекции {last_collection}') 
   LinksGoods = []  
   for LinkPage in LinkPages: 
        driver.get(LinkPage)
        try: 
            ListLG = driver.find_elements(By.XPATH,"//h2[@class = 'product-name']/a")
            for LG in ListLG: LinksGoods.append(LG.get_attribute("href")) 
        except: print("Категория не имеет товаров")      
   ParsingGoods(driver, LinksGoods,last_collection)    
   # LinksGoods.clear()

def ParsingGoods(driver, LinksGoods, last_collection):
    
    Goods = []  
    for Link in LinksGoods: 
        print(last_collection)
        driver.get(Link)   
        try:
           # Отбор товаров  
           OnSale = driver.find_element(By.XPATH,"//span[@class = 'in-stock']").text 
           if OnSale  == "В наличии":
                # Cбор данных товаров
                Article = driver.find_element(By.XPATH,"//span[@class = 'sku']").text
                Color = Article.split('/')[-1]
    
                FName = driver.find_element(By.XPATH,"//h1").text 
                Name1 = FName + " "+ Color
                Brand = FName.split()[-1]

                try:
                    Price = driver.find_element(By.XPATH,"//div[@class = 'product-price with-discount']/span[@itemprop = 'price']").text
                except: Price = driver.find_element(By.XPATH,"//div[@class = 'product-price']/span[@itemprop = 'price']").text
    
                SizeList =[]
                Sizes = driver.find_elements(By.XPATH,"//div[@class = 'product_order']/span")
                for Siz in Sizes:
                     SizeList.append(Siz.text) 
                     Size = ", ".join(SizeList) + "." 
         
                DescrTable = driver.find_elements(By.XPATH,"//div[@class ='product-fields']//strong[@itemprop ='value']")
                for Index, Discr in enumerate(DescrTable):
                    if Index == 2:
                       Season = "Сезон:" + Discr.text 
                    if Index == 4:
                       UpperMaterial = "Материал верха" + Discr.text 
                    if Index == 5:
                       LiningMaterial = "Материал подкладок" + Discr.text
                    if Index == 6:
                       InsoleMaterial = "Материал стелек" + Discr.text
               
                Description = driver.find_element(By.XPATH,"//div[@class= 'product-description']//span").text + " " + Season + " " + UpperMaterial +" "+ LiningMaterial + " " + InsoleMaterial
        
                # TableSize = []
                # try: 
                #     #Модальное окно не открывается при нажатии
                #     LinkTbSz = driver.find_element(By.XPATH,"//a[@id = 'tbsize-a']").get_attribute("href")
                #     LinkTbSz.click()
                
                #     TableSizes = driver.find_elements(By.XPATH,"//div[@id= 'tab1']//td")           
                #     for TS in TableSizes:                
                #         TableSize.append(TS.text)      
                #     indsize = TableSize.index("Размер")
                #     TabSZ = ", ".join(TableSize[:indsize + 1]) + "\n" + ", ".join(map(str, TableSize[indsize + 1:]))   
                # except:
                #     print("Таблица размеров не найдена")
                # print(TabSZ)   

                Picture = []
                Pictures = driver.find_elements(By.XPATH,"//div[@class ='img-container']/a")
                if len(Pictures) > 5:
                     for index, Pict in enumerate(Pictures):                      
                        if index % 2 != 0:
                             Picture.append(Pict.get_attribute("src"))
                else:
                    for Pict in Pictures: 
                        Picture.append(Pict.get_attribute("href"))          
 
                # Запись данных в экземляр структуры StructureOfProducts    
                StructureOfProduct = {
                     TabInd.NAME : Name1,
                     TabInd.ARTICLE : Article,
                     TabInd.BRAND : Brand,
                     TabInd.PRICE : Price, 
                     TabInd.SIZE :Size,
                     TabInd.DESCRIPTION : Description,
                     TabInd.PHOTO : Picture,
                     TabInd.LINK : Link,
                     }  
                Goods.append(StructureOfProduct)        
        except:
            print("При сборе данных возникла ошибка") 
    RecordingToExcel(Goods,last_collection)       
    # Goods.clear()
    
def RecordingToExcel(Goods,CollectionName): 
    
    if os.path.exists(file_name):
        wb = load_workbook(file_path)
        print(f"RecordingInExcel: Файл '{file_name}' успешно загружен.")    
        if CollectionName in wb.sheetnames:
            ws = wb[CollectionName] 
            print(f"RecordingInExcel: Лист '{CollectionName}' уже существует. Данные будут обновлены.")     
            if ws.cell(row=2, column=1).value is not None:
                for row in ws.iter_rows():
                    for cell in row: 
                        cell.value = None    
        else:
            ws = wb.create_sheet(title=CollectionName)
            print(f"Создан новый лист '{CollectionName}'.")
    else:
       wb = Workbook()  
       print(f"RecordingInExcel:Файл '{file_name}' успешно создан.")     
       ws = wb.create_sheet(title=CollectionName)
       print(f"RecordingInExcel:Создан новый лист '{CollectionName}'.")

    headers = ['Название', 'Артикл', 'Бренд', 'Цена', 'Размер', 'Описание', 'Изображение', 'Изображение1', 'Изображение2', 'Изображение3', 'Ссылка на товар']  
    for col_num, header in enumerate(headers, start=1):
         ws.cell(row=1, column=col_num, value=header)  
    for index, item in enumerate(Goods, start = 2 ):  
        ws.cell(row=index, column=1, value=item[TabInd.NAME])
        ws.cell(row=index, column=2, value=item[TabInd.ARTICLE])
        ws.cell(row=index, column=3, value=item[TabInd.BRAND])
        ws.cell(row=index, column=4, value=item[TabInd.PRICE])
        ws.cell(row=index, column=5, value=item[TabInd.SIZE])
        ws.cell(row=index, column=6, value=item[TabInd.DESCRIPTION])
        for imgindex, img in zip(range(4), item[TabInd.PHOTO]):   
             ws.cell(row=index, column=10 - imgindex, value = img) 
        ws.cell(row=index, column=11, value=item[TabInd.LINK ])
    wb.save(file_path)
    print(f"RecordingInExcel: Данные по категории {CollectionName} успешно записаны в файл '{file_name}'.")   
    
    try:  
        wb["Sheet"].title = "Лист1"
        wb.remove(wb["Лист1"])     
        wb.save(file_path)   
    except: pass 

# Главнвя процедура
s=Service(ChromedriverPuth)
driver = webdriver.Chrome(service=s)

GettingColltction(driver)
print()

#То что не вышло:
  # ModlWind = driver.find_element(By.XPATH,"//div[@class = 'navbar-header col-md-1']/button")
    # ModlWind.click()
   
    # MenyCats = driver.find_elements(By.XPATH,"//ul[@class = 'nav-child--show']/li/a")
    # for Index, MenyCat in enumerate(MenyCats):
    #     if Index != 3:
    #         Insetlinks = MenyCat.find_elements(By.TAG_NAME, "li")
    #         for Index, MenyCat in enumerate(MenyCats):
    #              if Index != 2:  
    #                  Catlinks = Insetlinks.find_elements(By.TAG_NAME, "a")
    #                  for CT in Catlinks:
    #                      Category = CT.text
    #                      CatLink = CT.get_attribute("href")
    #                      driver.get(CatLink)
    #//ul[@class = 'nav-child unstyled small']/li
    
    # WBootsCat = {    
    #     "Кросовики": "Жен Кросовики Кеды", 
    #     "Кеды": "Жен Кросовики Кеды",
    #     "Ботинки": "Жен Ботинки Ботильоны",  
    #     "Ботильоны": "Жен Ботинки Ботильоны", 
    #     "Сапоги":"Жен Сапоги Полусапоги", 
    #     "Полусапоги": "Жен Сапоги Полусапоги",
    #     "Туфли": "Жен Туфли", 
    #     "Босоножки": "Жен Босоножки Сандали", 
    #     "Сандали": "Жен Босоножки Сандали",
    #     "Балентки": "Жен Балетки Макасины", 
    #     "Макасины" : "Жен Балетки Макасины",
    #     "Сабо" : "Жен Сабо Сланцы", 
    #     "Сланцы" : "Жен Сабо Сланцы",
    #     "Дутики" : "Жен Дутики Валенки", 
    #     "Валенки" : "Жен Дутики Валенки",
    #     "Демисезон" : "Жен Демисезон",
    #     "ЕвроЗима" : "Жен ЕвроЗима",
    #     "Зима" : "Жен Зима",  
    #     "Лето" : "Жен Лето"
    # }
    
    # driver.get("https://www.okeanobuvi.ru/")
    # time.sleep(5)
    
    # #Обработка появления меню на сайте
    # actions = ActionChains(driver)
    # Menu = driver.find_element(By.XPATH,"//a[@class = ' dropdown-toggle'][1]")
    # actions.move_to_element(Menu).perform()
    # #Сбор категорий женской одежды
    # РrocessedLink = set()
    # VCats = driver.find_elements(By.XPATH,"//ul[@class = 'mega-nav level2']/li/a") 
    # LinksCol = []    
    # for i in range(min(19, len(VCats))):
    #     ct = VCats[i]
    #     СatName = ct.text  
    #     if СatName  in WBootsCat:
    #         NameCol = WBootsCat[СatName]
    #     CatLink = ct.get_attribute("href")
       
    #     if CatLink  not in РrocessedLink:
    #         РrocessedLink.add(CatLink) 
    #         LinksCol.append(CatLink)

    # driver.get("https://www.okeanobuvi.ru/%D0%BE%D0%B1%D1%83%D0%B2%D1%8C/%D0%B6%D0%B5%D0%BD%D1%81%D0%BA%D0%B0%D1%8F/%D0%B1%D0%B0%D0%BB%D0%B5%D1%82%D0%BA%D0%B8")
    # time.sleep(10)     